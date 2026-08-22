"""
Script untuk mengimpor data anggota dari DATA IT.xlsx ke Supabase.
Jalankan dari folder absensi-it-26/:
  python import_anggota.py

Requirements:
  pip install openpyxl requests
"""

import openpyxl
import requests
import os
import sys
import json

# Fix encoding Windows terminal
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# ─── KONFIGURASI ──────────────────────────────────────────────
# Isi dengan data dari Supabase Project Settings > API
SUPABASE_URL = "https://vomaluikqvcryocefoke.supabase.co"
SUPABASE_SERVICE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InZvbWFsdWlrcXZjcnlvY2Vmb2tlIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc4NzM2MDU5NiwiZXhwIjoyMTAyOTM2NTk2fQ.H6ZLhWGF79QSkLGiRjMsOoBuixkCp0cXeXQ2_7I18PA"
# ──────────────────────────────────────────────────────────────

# Cari file Excel: coba di folder saat ini, lalu folder parent
EXCEL_CANDIDATES = [
    "DATA IT.xlsx",
    "../DATA IT.xlsx",
    os.path.join(os.path.dirname(__file__), "DATA IT.xlsx"),
    os.path.join(os.path.dirname(__file__), "..", "DATA IT.xlsx"),
]

def find_excel():
    for path in EXCEL_CANDIDATES:
        if os.path.exists(path):
            return os.path.abspath(path)
    return None

def main():
    excel_path = find_excel()
    if not excel_path:
        print("❌ File 'DATA IT.xlsx' tidak ditemukan.")
        print("   Letakkan file Excel di folder yang sama dengan script ini, atau di folder induknya.")
        sys.exit(1)

    print(f"📂 Membaca file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    print(f"📋 Kolom ditemukan: {headers}")

    # Cari indeks kolom (case-insensitive)
    headers_lower = [str(h).strip().lower() if h else "" for h in headers]
    try:
        idx_nrp   = headers_lower.index("nrp")
        idx_nama  = headers_lower.index("nama")
        idx_prodi = headers_lower.index("program studi")
    except ValueError as e:
        print(f"❌ Kolom tidak ditemukan: {e}")
        print(f"   Kolom yang ada: {headers}")
        sys.exit(1)

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    anggota_list = []
    for row in rows:
        nrp   = str(row[idx_nrp]).strip()   if row[idx_nrp]   else None
        nama  = str(row[idx_nama]).strip()  if row[idx_nama]  else None
        prodi = str(row[idx_prodi]).strip() if row[idx_prodi] else None

        if nrp and nama and prodi and nrp.lower() != "none":
            anggota_list.append({"nrp": nrp, "nama": nama, "program_studi": prodi})

    print(f"📊 {len(anggota_list)} anggota siap diimport...\n")

    # HTTP session dengan requests (lebih stabil dari httpx)
    session = requests.Session()
    session.headers.update({
        "apikey": SUPABASE_SERVICE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates",
    })

    batch_size = 20  # Kecil agar tidak di-reject server
    success = 0
    failed = 0

    for i in range(0, len(anggota_list), batch_size):
        batch = anggota_list[i:i + batch_size]
        batch_num = i // batch_size + 1

        try:
            resp = session.post(
                f"{SUPABASE_URL}/rest/v1/anggota",
                data=json.dumps(batch),
                timeout=30,
                verify=True,
            )

            if resp.status_code in (200, 201):
                success += len(batch)
                print(f"  ✅ Batch {batch_num}: {len(batch)} anggota berhasil ({i+1}–{i+len(batch)})")
            else:
                failed += len(batch)
                print(f"  ❌ Batch {batch_num} gagal [{resp.status_code}]: {resp.text[:200]}")

        except requests.exceptions.SSLError as e:
            print(f"  ⚠️  Batch {batch_num}: SSL error, mencoba tanpa verifikasi...")
            try:
                resp = session.post(
                    f"{SUPABASE_URL}/rest/v1/anggota",
                    data=json.dumps(batch),
                    timeout=30,
                    verify=False,
                )
                if resp.status_code in (200, 201):
                    success += len(batch)
                    print(f"  ✅ Batch {batch_num}: {len(batch)} anggota berhasil (no-SSL-verify)")
                else:
                    failed += len(batch)
                    print(f"  ❌ Batch {batch_num} gagal [{resp.status_code}]: {resp.text[:200]}")
            except Exception as e2:
                failed += len(batch)
                print(f"  ❌ Batch {batch_num} error: {e2}")

        except Exception as e:
            failed += len(batch)
            print(f"  ❌ Batch {batch_num} error: {type(e).__name__}: {e}")

    print(f"\n{'='*50}")
    print(f"🎉 Selesai!")
    print(f"   ✅ Berhasil : {success} anggota")
    if failed:
        print(f"   ❌ Gagal   : {failed} anggota")
    print(f"{'='*50}")

if __name__ == "__main__":
    main()
